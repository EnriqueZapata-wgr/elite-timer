# -*- coding: utf-8 -*-
"""Genera el cuadernillo de revision clinica de la matriz (13 decisiones).

Entregable de una sola pasada. No lee ni modifica la matriz: todo el contenido
viene del documento R and D/MATRIZ_V7_V6_10_PENDIENTES_PARA_FIRMA.md
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


OUT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "R and D",
    "Revision_clinica_matriz_13_decisiones_ago2026.xlsx",
)

# ---------------------------------------------------------------- paleta
AZUL = "1F3B54"        # encabezados nuestros
AMBAR = "B8860B"       # encabezados de ella
AMARILLO = "FFF6D5"    # celdas que ella llena
GRIS = "F2F2F2"
BLANCO = "FFFFFF"
ROJO_TXT = "9C1F1F"
AMBAR_TXT = "8A5A00"
VERDE_TXT = "1F5C2E"

F = "Arial"
thin = Side(style="thin", color="BFBFBF")
BORDE = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ============================================================ HOJA 1
h1 = wb.active
h1.title = "Empieza aquí"
h1.sheet_view.showGridLines = False

h1.column_dimensions["A"].width = 2
h1.column_dimensions["B"].width = 26
h1.column_dimensions["C"].width = 95

def bloque(row, titulo, texto, alto=None):
    h1.cell(row=row, column=2, value=titulo).font = Font(name=F, size=11, bold=True, color=AZUL)
    h1.cell(row=row, column=2).alignment = Alignment(vertical="top", wrap_text=True)
    c = h1.cell(row=row, column=3, value=texto)
    c.font = Font(name=F, size=11)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    if alto:
        h1.row_dimensions[row].height = alto

h1["B2"] = "Revisión clínica de la matriz de salud funcional"
h1["B2"].font = Font(name=F, size=18, bold=True, color=AZUL)
h1.row_dimensions[2].height = 26

h1["B3"] = "13 decisiones pendientes · agosto de 2026"
h1["B3"].font = Font(name=F, size=12, color="666666")
h1.row_dimensions[3].height = 20
h1.row_dimensions[4].height = 8

bloque(5, "Qué es esto",
       "La matriz que califica los resultados de laboratorio dentro de la app tiene 13 puntos "
       "que no podemos resolver sin criterio clínico. Aquí están, uno por renglón.", 32)
bloque(6, "Por qué te llega a ti",
       "La matriz es tuya y el criterio clínico lo firmas tú. No cambiamos ningún valor sin tu respuesta.", 32)
bloque(7, "Qué necesitamos",
       "Abre la hoja \"Los 13 casos\". En cada renglón hay una pregunta directa. "
       "Contéstala en las tres columnas amarillas del final: tu respuesta, comentarios y la fecha en que lo revisaste. "
       "Nada más que eso.", 46)
bloque(8, "Cuánto te va a tomar",
       "Entre 30 y 45 minutos. Cinco de los 13 son sólo confirmar un sí, y ya te dejamos escrito lo que creemos.", 32)
bloque(9, "Para cuándo",
       "Antes del 25 de agosto de 2026. El lanzamiento es el 1 de septiembre y necesitamos margen "
       "para aplicar los cambios y probarlos.", 32)
bloque(10, "Si algo no se entiende",
       "Escríbelo en Comentarios y lo resolvemos. Una respuesta parcial sirve mucho más que ninguna.", 32)
bloque(11, "La tercera hoja",
       "\"Detalle técnico\" tiene las cifras exactas de cada caso, por si quieres verlas. "
       "No hace falta abrirla para contestar.", 32)

h1.row_dimensions[12].height = 10

# leyenda + ejemplo
lbl = h1.cell(row=13, column=2, value="Así se llena")
lbl.font = Font(name=F, size=11, bold=True, color=AMBAR_TXT)
lbl.alignment = Alignment(vertical="top")
ej = ("Sólo escribes en las tres columnas amarillas. Ejemplo, para el renglón de la insulina:\n"
      "   Tu respuesta:  Sí, µUI/mL\n"
      "   Comentarios:   Es la unidad que reportan todos los laboratorios con los que trabajamos.\n"
      "   Fecha:         18/08/2026")
c = h1.cell(row=13, column=3, value=ej)
c.font = Font(name=F, size=10.5)
c.alignment = Alignment(wrap_text=True, vertical="top")
c.fill = PatternFill("solid", fgColor=AMARILLO)
c.border = BORDE
h1.row_dimensions[13].height = 70

h1.row_dimensions[14].height = 10
n = h1.cell(row=15, column=3,
            value="No se modificó ningún valor de la matriz. Este archivo sólo describe lo que hay hoy "
                  "y lo que produce cuando la app lo usa.")
n.font = Font(name=F, size=9, italic=True, color="808080")
n.alignment = Alignment(wrap_text=True, vertical="top")

h1.print_area = "B1:C16"
h1.page_setup.orientation = "portrait"
h1.page_setup.fitToWidth = 1
h1.page_setup.fitToHeight = 1
h1.sheet_properties.pageSetUpPr.fitToPage = True

# ============================================================ HOJA 2
h2 = wb.create_sheet("Los 13 casos")
h2.sheet_view.showGridLines = False

ENC = ["No.", "Urgencia", "Parámetro", "Qué pasa hoy", "Ejemplo con números",
       "Qué necesitamos de ti", "TU RESPUESTA", "COMENTARIOS", "FECHA DE REVISIÓN"]
ANCHOS = [5, 14, 24, 44, 52, 50, 30, 30, 18]

CASOS = [
    # (urgencia, parametro, que pasa, ejemplo, pregunta)
    ("Urgente", "T3 libre",
     "El rango óptimo de T3 libre está etiquetado en ng/dL, pero las cifras son de pg/mL. "
     "Cuando el laboratorio reporta en pmol/L, el número cae dentro del rango por pura coincidencia.",
     "Una T3 libre de 3.5 pmol/L, que es el piso del rango de referencia y hace sospechar conversión "
     "periférica pobre, el sistema la califica 100 de 100 y le dice a la persona \"estás produciendo y "
     "convirtiendo bien\". Ese valor equivale a 2.28 pg/mL y le correspondería 25 de 100 y \"pide atención\".",
     "¿Cuál es la unidad de referencia de la T3 libre y su rango óptimo en esa unidad? "
     "Creemos que las cifras 3.2 a 4.2 son pg/mL y lo único mal es la etiqueta. ¿Lo confirmas o prefieres otro rango?"),

    ("Urgente", "Apolipoproteína B",
     "Uno de los valores límite de la ApoB quedó en cero y desordena la escala. "
     "El resultado es que cualquier ApoB entre 0 y 39 mg/dL se califica como aceptable.",
     "Una ApoB de 15 mg/dL, que obliga a descartar malabsorción o hipobetalipoproteinemia, "
     "se califica 50 de 100, se muestra como \"Aceptable\" y viene acompañada del texto "
     "\"menos partículas circulando se lee bien\".",
     "¿A partir de qué cifra en mg/dL una ApoB baja deja de ser plausible y debe marcarse? "
     "Creemos que el valor va entre 30 y 40. ¿Qué número pones?"),

    ("Urgente", "Testosterona total (lectura de sueño)",
     "La testosterona total se evalúa dos veces, una desde el sistema hormonal y otra desde el sueño. "
     "La lectura de sueño usa cifras masculinas para los dos sexos y está etiquetada en la unidad equivocada.",
     "Un hombre con testosterona total de 18 ng/dL, nivel de castración, sale \"Aceptable\" en la lectura de sueño. "
     "Una mujer con 40 ng/dL, normal, sale correcta en la hormonal pero se le descuenta la edad de sueño "
     "con 0 de 100, sin que la pantalla se lo explique.",
     "¿La testosterona debe leerse con el mismo rango en sueño que en sistema hormonal, cada sexo con el suyo? "
     "Si tiene criterio propio, ¿cuál es el rango de sueño para mujeres y cuál para hombres, y en qué unidad?"),

    ("Alta", "LDH en mujeres (lectura de inflamación)",
     "El rango de LDH para mujeres quedó copiado del índice neutrófilos sobre linfocitos (0.1 a 1.5), "
     "así que ninguna LDH real cabe ahí.",
     "Una mujer con LDH de 180 U/L, perfectamente normal, ve su renglón en rojo con la leyenda \"Pide atención\" "
     "y al lado impreso \"rango 0.1 a 1.5\". Le pasa al 100% de las mujeres: den 120, 180 o 400, todas sacan 0 de 100.",
     "¿Qué rango de LDH en U/L aplica para mujeres? Creemos que el mismo que en hombres, 167 a 187 U/L. "
     "¿Lo confirmas o prefieres otro?"),

    ("Alta", "Testosterona total en mujeres (umbral de lectura)",
     "La app traduce el resultado de ng/dL a ng/mL sólo cuando pasa de 20. Ese umbral parte a la mitad "
     "el rango femenino normal, que va aproximadamente de 15 a 70 ng/dL.",
     "Una mujer con testosterona total de 18 ng/dL, que es bajo normal, se califica 0 de 100 y \"Pide atención\". "
     "Si el valor se tradujera bien, le tocaría 80 de 100 y \"Aceptable\".",
     "¿Por debajo de qué cifra de testosterona total en ng/dL ya no es plausible encontrar a una mujer? "
     "Ese número es el que necesitamos."),

    ("Alta", "Edad corporal contra edad cronológica",
     "La diferencia entre edad corporal y edad cronológica se evalúa en tres lugares con rangos distintos, "
     "y el de sueño tiene tres ceros de más al inicio que desordenan la escala.",
     "Alguien con edad corporal 12 años por debajo de su edad cronológica, o sea un resultado excelente, "
     "saca 100 de 100 por composición corporal y al mismo tiempo 0 de 100 con \"Pide atención\" por sueño y por vitalidad. "
     "Las tres cosas en la misma pantalla.",
     "¿Qué tan por debajo de la edad cronológica sigue contando como bueno? "
     "Creemos que el mismo rango en los tres lugares, de 15 años por debajo hasta 1 año por debajo. ¿Lo confirmas?"),

    ("Media", "Ácido úrico en hombres",
     "En hombres el ácido úrico se evalúa con dos rangos distintos: 4 a 6 mg/dL cuando se lee como inflamación "
     "y 3.5 a 5.5 mg/dL cuando se lee como función renal. En mujeres los dos coinciden y no hay problema.",
     "Un hombre con ácido úrico de 5.8 mg/dL ve \"En tu ventana\" en pantalla mientras su edad renal se descuenta "
     "por detrás. Con 3.8 mg/dL pasa lo contrario: la pantalla le exige de más sobre un valor que la lectura renal "
     "considera óptimo.",
     "¿La lectura inflamatoria del ácido úrico es distinta de la renal, a propósito? "
     "Si es un solo criterio, ¿cuál de los dos rangos manda en hombres?"),

    ("Media", "LDH, un criterio o dos",
     "La LDH tiene dos rangos muy distintos según desde dónde se lea: 167 a 187 U/L como inflamación "
     "y 20 a 200 U/L como inmunidad. Uno es casi diez veces más ancho que el otro.",
     "Un hombre con LDH de 250 U/L sale \"Pide atención\" por un lado y \"Aceptable\" por el otro, al mismo tiempo. "
     "Con 150 U/L sale \"Aceptable\" y \"En tu ventana\".",
     "¿La LDH tiene un solo criterio o dos según se lea como inflamación o como inmunidad? "
     "Si es uno solo, ¿cuál rango en U/L? Y de paso: ¿167 a 187 U/L te parece correcto o demasiado estrecho?"),

    ("Media", "T3 libre por arriba del rango",
     "La T3 libre no tiene graduación hacia arriba. Cualquier valor por encima del rango óptimo "
     "cae de golpe en la peor calificación, sin escalones intermedios.",
     "Una T3 libre apenas por encima del techo y una francamente alta reciben las dos exactamente "
     "la misma calificación de 0 de 100 y la misma etiqueta.",
     "¿Quieres escalones hacia arriba en la T3 libre? Si sí, ¿qué tres cifras marcan el ascenso? "
     "Si prefieres que quede como está, dilo y lo dejamos escrito."),

    ("Sólo confirmar", "LDH, unidad",
     "La LDH aparece etiquetada como \"Ratio\" cuando se mide en unidades por litro. "
     "Las cifras están bien, la etiqueta no.",
     "En pantalla la persona lee \"LDH 180 Ratio\". El número es correcto, la unidad no corresponde a nada.",
     "Creemos que la unidad correcta es U/L. ¿Lo confirmas?"),

    ("Sólo confirmar", "Homocisteína, unidad",
     "La homocisteína está etiquetada como \"mcmol/ml\", que no existe como unidad. "
     "Las cifras del rango sí son correctas y la calificación sale bien.",
     "Una persona con homocisteína de 7 lee en pantalla \"7 mcmol/ml, rango 5 a 9 mcmol/ml\". "
     "El número está bien y la unidad está inventada.",
     "Creemos que la unidad correcta es µmol/L. ¿Lo confirmas?"),

    ("Sólo confirmar", "Insulina, unidad",
     "La insulina está etiquetada como \"mgUI/ml\", que tampoco existe como unidad. "
     "Las cifras del rango son correctas y la calificación sale bien.",
     "Una persona con insulina de 4 lee en pantalla \"4 mgUI/ml, rango 2 a 6 mgUI/ml\".",
     "Creemos que la unidad correcta es µUI/mL. ¿Lo confirmas?"),

    ("Sólo confirmar", "HbA1c, hematocrito y RDW-CV",
     "Los tres están escritos en el archivo como fracción decimal (0.049) y etiquetados en porcentaje. "
     "La app los traduce antes de mostrarlos, así que en pantalla se ven bien. "
     "Lo confuso vive dentro del archivo de la matriz.",
     "Una HbA1c de 5.1 % se guarda como 0.051, se compara contra 0.049 a 0.052 y se muestra correctamente "
     "como \"5.1 %, rango 4.9 a 5.2 %\". Todo bien para la persona, ilegible para quien abra la matriz.",
     "Creemos que conviene reescribir los tres en porcentaje (4.9 en lugar de 0.049) para que la matriz "
     "se entienda sola, sin cambiar ninguna calificación. ¿Lo confirmas o prefieres dejarlos en fracción con una nota?"),
]

COL_URG = {"Urgente": ROJO_TXT, "Alta": AMBAR_TXT, "Media": AZUL, "Sólo confirmar": VERDE_TXT}

# titulo
h2.merge_cells("A1:I1")
h2["A1"] = "Los 13 casos · escribe sólo en las tres columnas amarillas"
h2["A1"].font = Font(name=F, size=13, bold=True, color=AZUL)
h2["A1"].alignment = Alignment(vertical="center")
h2.row_dimensions[1].height = 24

HDR = 2
for i, (t, w) in enumerate(zip(ENC, ANCHOS), start=1):
    h2.column_dimensions[get_column_letter(i)].width = w
    c = h2.cell(row=HDR, column=i, value=t)
    c.fill = PatternFill("solid", fgColor=AMBAR if i >= 7 else AZUL)
    c.font = Font(name=F, size=10.5, bold=True, color=BLANCO)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDE
h2.row_dimensions[HDR].height = 34


def alto_estimado(valores):
    lineas = 1
    for idx, texto in enumerate(valores):
        if not isinstance(texto, str):
            continue
        ancho = ANCHOS[idx]
        chars = max(1, int(ancho * 0.95))
        n = 0
        for parrafo in texto.split("\n"):
            n += max(1, -(-len(parrafo) // chars))
        lineas = max(lineas, n)
    return max(30, lineas * 14 + 8)


fila = HDR + 1
# renglon de ejemplo
ejemplo = ["EJ.", "Sólo confirmar", "Ejemplo, borra este renglón",
           "Así se ve un caso descrito.", "Así se ve el ejemplo con cifras.",
           "Así se ve la pregunta.", "Sí, µUI/mL",
           "Es la unidad que reportan los laboratorios con los que trabajamos.", "18/08/2026"]
for i, v in enumerate(ejemplo, start=1):
    c = h2.cell(row=fila, column=i, value=v)
    c.font = Font(name=F, size=9.5, italic=True, color="909090")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.fill = PatternFill("solid", fgColor=AMARILLO if i >= 7 else GRIS)
    c.border = BORDE
h2.row_dimensions[fila].height = 38
fila += 1

primera_datos = fila
for n, (urg, param, hoy, ej, preg) in enumerate(CASOS, start=1):
    vals = [n, urg, param, hoy, ej, preg, None, None, None]
    for i, v in enumerate(vals, start=1):
        c = h2.cell(row=fila, column=i, value=v)
        c.border = BORDE
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal="center" if i in (1, 2, 9) else "left")
        if i == 1:
            c.font = Font(name=F, size=11, bold=True, color=AZUL)
        elif i == 2:
            c.font = Font(name=F, size=10, bold=True, color=COL_URG[urg])
        elif i == 3:
            c.font = Font(name=F, size=10.5, bold=True)
        elif i <= 6:
            c.font = Font(name=F, size=10)
        else:
            c.font = Font(name=F, size=10.5)
            c.fill = PatternFill("solid", fgColor=AMARILLO)
    
    h2.row_dimensions[fila].height = alto_estimado([str(v) if v else "" for v in vals])
    fila += 1
ultima_datos = fila - 1

for r in range(primera_datos, ultima_datos + 1):
    h2.cell(row=r, column=9).number_format = "DD/MM/YYYY"

pie = h2.cell(row=ultima_datos + 2, column=3,
              value="Si algo no se entiende o falta un dato para decidir, escríbelo en Comentarios. "
                    "Una respuesta parcial sirve más que ninguna.")
pie.font = Font(name=F, size=10, italic=True, color="808080")
h2.merge_cells(start_row=ultima_datos + 2, start_column=3, end_row=ultima_datos + 2, end_column=6)

h2.freeze_panes = "D3"
h2.auto_filter.ref = "A2:I%d" % ultima_datos
h2.print_title_rows = "1:2"
h2.print_area = "A1:I%d" % (ultima_datos + 2)
h2.page_setup.orientation = "landscape"
h2.page_setup.paperSize = h2.PAPERSIZE_A4
h2.page_setup.fitToWidth = 1
h2.page_setup.fitToHeight = 0
h2.sheet_properties.pageSetUpPr.fitToPage = True

# ============================================================ HOJA 3
h3 = wb.create_sheet("Detalle técnico")
h3.sheet_view.showGridLines = False

ENC3 = ["Casos", "Parámetro", "Sexo", "Se lee desde", "Etiqueta de unidad en la matriz",
        "Unidad real del resultado", "Los ocho valores límite tal como están escritos hoy",
        "Rango óptimo", "Peso", "Qué produce"]
ANCHOS3 = [8, 22, 12, 22, 20, 20, 44, 16, 8, 58]

VACIO = "(vacía)"
DET = [
    ("1, 9", "T3 libre", "Ambos", "Sistema hormonal", "ng/dl", "pg/mL",
     "2.2 · 2.5 · 2.8 · 3.2 · 4.2 · %s · %s · %s" % (VACIO, VACIO, VACIO), "3.2 a 4.2", "0.10",
     "En ng/dL la T3 libre normal ronda 0.2 a 0.5, un orden de magnitud por debajo de lo escrito. "
     "Las cifras corresponden a pg/mL, que es lo que guarda la base. El riesgo real es la coincidencia con pmol/L: "
     "3.5 a 6.5 pmol/L equivale a 2.28 a 4.23 pg/mL (1 pmol/L = 0.651 pg/mL), casi encima del intervalo escrito "
     "pero desplazado. Los tres valores superiores vacíos dejan sin escalones el exceso."),

    ("2", "Apolipoproteína B", "Ambos", "Cardiovascular", "mg/dl (correcta)", "mg/dL",
     "30 · 0 · 40 · 50 · 99 · 110 · 125 · 150", "50 a 99 mg/dL", "0.15",
     "El segundo valor es cero y rompe el orden ascendente entre el 30 y el 40. Efecto uno: el tramo de 25 puntos "
     "pide \"mayor o igual a 30 y menor que 0\", que es imposible, y nadie puede sacar 25. Efecto dos, el que hace daño: "
     "el tramo de 50 puntos pasa a pedir \"mayor o igual a 0 y menor que 40\" y absorbe todo el extremo bajo. "
     "Es el peso más alto de todo el cardiovascular."),

    ("3", "Testosterona total", "Hombres", "Sistema hormonal", "ng/ml", "ng/dL en la base",
     "7 a 12 es el rango óptimo; los intermedios no cambian esta decisión", "7 a 12 ng/mL", "0.12",
     "Funciona bien. La app divide entre 100 los valores por arriba de 20 para poder compararlos, y eso ya está en producción."),
    ("3", "Testosterona total", "Hombres", "Sueño", "ng/dl", "ng/dL en la base",
     "3 · 4 · 5 · 7 · 13 · 15 · 20 · 25", "7 a 13", "0.07",
     "Las cifras son magnitudes de ng/mL con etiqueta de ng/dL. En ng/dL de verdad, 13 sería testosterona de castración. "
     "Una testosterona de 18 ng/dL no se traduce (el umbral es 20), cae entre 15 y 20 y saca 50 de 100 con etiqueta \"Aceptable\"."),
    ("3", "Testosterona total", "Mujeres", "Sistema hormonal", "ng/ml", "ng/dL en la base",
     "0.2 a 0.55 es el rango óptimo", "0.2 a 0.55 ng/mL", "0.05",
     "Correcto. Equivale a 20 a 55 ng/dL, rango femenino razonable."),
    ("3", "Testosterona total", "Mujeres", "Sueño", "ng/dl", "ng/dL en la base",
     "3 · 4 · 5 · 7 · 13 · 15 · 20 · 25", "7 a 13", "0.07",
     "Repite el rango masculino dentro del sexo femenino. 7 a 13 equivale a 700 a 1300 ng/dL. "
     "La diferencia contra el rango hormonal del mismo sexo es de factor 23 a 35 según el extremo. "
     "Toda mujer que suba una testosterona total recibe 0 en la edad de sueño."),

    ("5", "Testosterona total", "Mujeres", "Traducción de unidades", "no aplica", "ng/dL en la base",
     "Umbral actual de traducción: 20", "no aplica", "no aplica",
     "El umbral de 20 se calibró sobre magnitudes masculinas. El rango femenino real va de 15 a 70 ng/dL, "
     "así que los valores de 15 a 20 ng/dL no se traducen y se comparan como si fueran ng/mL. "
     "18 ng/dL saca 0 de 100 cuando le tocaría 80."),

    ("4, 8, 10", "LDH", "Hombres", "Inflamación", "Ratio", "U/L",
     "109 · 120 · 135 · 167 · 187 · 205 · 220 · 246", "167 a 187 U/L", "0.05",
     "Rango de 20 U/L de ancho contra 180 U/L en la lectura de inmunidad del mismo parámetro y el mismo sexo. "
     "No es un error de captura evidente, pero es una diferencia de criterio de casi un orden de magnitud."),
    ("4, 8, 10", "LDH", "Hombres", "Inmunidad", "Ratio", "U/L",
     "5 · 10 · 15 · 20 · 200 · 300 · 400 · 500", "20 a 200 U/L", "0.05",
     "150 U/L sale \"Aceptable\" por inflamación y \"En tu ventana\" por inmunidad. 250 U/L sale \"Pide atención\" y \"Aceptable\"."),
    ("4, 8, 10", "LDH", "Mujeres", "Inflamación", "Ratio", "U/L",
     "%s · %s · %s · 0.1 · 1.5 · 2 · 2.2 · 2.5" % (VACIO, VACIO, VACIO), "0.1 a 1.5", "0.05",
     "Es idéntica renglón por renglón a la fila del índice neutrófilos sobre linfocitos, tres filas arriba en el mismo lugar. "
     "Es una copia de celda, no un criterio. Como la LDH se mide en U/L y no baja de 100 en la práctica, "
     "ningún valor real cabe: el 100% de las mujeres saca 0 y ve \"Pide atención\"."),
    ("4, 8, 10", "LDH", "Mujeres", "Inmunidad", "Ratio", "U/L",
     "5 · 10 · 15 · 20 · 200 · 300 · 400 · 500", "20 a 200 U/L", "0.05",
     "Esta sí califica bien, pero no es la que se muestra en pantalla."),

    ("7", "Ácido úrico", "Hombres", "Inflamación", "mg/dl", "mg/dL",
     "%s · %s · 3 · 4 · 6 · 7 · 8 · 9" % (VACIO, VACIO), "4 a 6 mg/dL", "0.06",
     "Es la que se muestra en pantalla, porque la inflamación va antes en el orden del archivo."),
    ("7", "Ácido úrico", "Hombres", "Renal y micronutrientes", "mg/dl", "mg/dL",
     "%s · %s · 3 · 3.5 · 5.5 · 6.5 · 7.5 · 8" % (VACIO, VACIO), "3.5 a 5.5 mg/dL", "0.10",
     "Corrida medio punto hacia abajo respecto a la de inflamación en los dos extremos, y pesa casi el doble. "
     "5.8 mg/dL da 100 por un lado y 80 por el otro; 3.8 mg/dL da 80 y 100. En mujeres los dos rangos coinciden."),

    ("6", "Edad corporal menos edad cronológica", "Ambos", "Sueño", "años", "años",
     "0 · 0 · 0 · -10 · -1 · 0 · 5 · 10", "-10 a -1", "según sub edad",
     "Los tres ceros iniciales están donde los otros dos lugares tienen celdas vacías, y dejan un tramo inalcanzable "
     "igual que en la ApoB. Un número negativo es lo bueno: mide qué tanto la edad corporal va por debajo de la cronológica."),
    ("6", "Edad corporal menos edad cronológica", "Ambos", "Composición corporal", "años", "años",
     "%s · %s · %s · -15 · -1 · 0 · 5 · 10" % (VACIO, VACIO, VACIO), "-15 a -1", "según sub edad",
     "Con 12 años por debajo de la cronológica da 100 de 100."),
    ("6", "Edad corporal menos edad cronológica", "Ambos", "Vitalidad", "años", "años",
     "%s · %s · %s · -10 · -1 · 0 · 5 · 10" % (VACIO, VACIO, VACIO), "-10 a -1", "según sub edad",
     "Con 12 años por debajo de la cronológica da 0 de 100 y \"Pide atención\". Contradice a composición corporal "
     "en la misma pantalla. No aparece en la lista de laboratorios porque no es un parámetro de laboratorio."),

    ("11", "Homocisteína", "Ambos", "Inflamación", "mcmol/ml", "µmol/L",
     "1 · 2 · 4 · 5 · 9 · 12 · 14 · 17", "5 a 9", "0.15",
     "La etiqueta no existe como unidad. Las cifras sí son µmol/L y la calificación sale correcta. "
     "La etiqueta se imprime literal en pantalla. Peso más alto de inflamación junto con la PCR."),
    ("12", "Insulina", "Ambos", "Metabolismo", "mgUI/ml", "µUI/mL",
     "%s · %s · %s · 2 · 6 · 8 · 10 · 15" % (VACIO, VACIO, VACIO), "2 a 6", "0.07",
     "Mismo caso. La etiqueta no existe, las cifras son µUI/mL, la calificación sale correcta "
     "y la etiqueta se imprime literal en pantalla."),

    ("13", "HbA1c", "Ambos", "Metabolismo", "%", "porcentaje",
     "0.01 · 0.025 · 0.035 · 0.049 · 0.052 · 0.056 · 0.058 · 0.06", "4.9 % a 5.2 %", "0.14",
     "Escrita en fracción decimal con etiqueta de porcentaje. La app la traduce antes de mostrarla, "
     "así que la pantalla es correcta. La confusión vive dentro del archivo. Peso más alto del metabolismo."),
    ("13", "Hematocrito", "Hombres", "Cardiovascular", "%", "porcentaje",
     "0.33 · 0.34 · 0.36 · 0.38 · 0.44 · 0.49 · 0.52 · 0.54", "38 % a 44 %", "0.02", "Mismo caso."),
    ("13", "Hematocrito", "Mujeres", "Cardiovascular", "%", "porcentaje",
     "0.32 · 0.321 · 0.36 · 0.38 · 0.43 · 0.45 · 0.48 · 0.481", "38 % a 43 %", "0.02",
     "Mismo caso. El techo femenino es 43 %, no 44 % como el masculino."),
    ("13", "RDW-CV", "Ambos", "Cardiovascular", "%", "porcentaje",
     "0.03 · 0.04 · 0.05 · 0.06 · 0.125 · 0.135 · 0.145 · 0.155", "6 % a 12.5 %", "0.04", "Mismo caso."),
]

h3.merge_cells("A1:J1")
h3["A1"] = "Detalle técnico de respaldo · no hace falta abrir esta hoja para contestar la anterior"
h3["A1"].font = Font(name=F, size=13, bold=True, color=AZUL)
h3["A1"].alignment = Alignment(vertical="center")
h3.row_dimensions[1].height = 24

for i, (t, w) in enumerate(zip(ENC3, ANCHOS3), start=1):
    h3.column_dimensions[get_column_letter(i)].width = w
    c = h3.cell(row=2, column=i, value=t)
    c.fill = PatternFill("solid", fgColor=AZUL)
    c.font = Font(name=F, size=10, bold=True, color=BLANCO)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDE
h3.row_dimensions[2].height = 40


def alto3(vals):
    lineas = 1
    for idx, texto in enumerate(vals):
        chars = max(1, int(ANCHOS3[idx] * 0.95))
        lineas = max(lineas, max(1, -(-len(str(texto)) // chars)))
    return max(28, lineas * 13.5 + 8)


r = 3
prev = None
for row in DET:
    banda = (prev is not None and row[1] != prev)
    for i, v in enumerate(row, start=1):
        c = h3.cell(row=r, column=i, value=v)
        c.border = BORDE
        c.font = Font(name=F, size=9.5, bold=(i == 2))
        c.alignment = Alignment(wrap_text=True, vertical="top",
                                horizontal="center" if i in (1, 3, 9) else "left")
        if i == 2 or banda:
            pass
    if (r % 2) == 1:
        for i in range(1, len(ENC3) + 1):
            h3.cell(row=r, column=i).fill = PatternFill("solid", fgColor=GRIS)
    h3.row_dimensions[r].height = alto3(row)
    prev = row[1]
    r += 1

nota = h3.cell(row=r + 1, column=1,
               value="Los ocho valores límite arman nueve tramos. Los dos centrales son el rango óptimo y valen 100. "
                     "Hacia afuera la calificación baja a 80, 50, 25 y 0. En la app, 100 se muestra como \"En tu ventana\", "
                     "80 y 50 como \"Aceptable\", 25 y 0 como \"Pide atención\". Cuando un mismo parámetro se lee desde dos "
                     "lugares, los dos entran al cálculo de la Edad ATP, pero la pantalla muestra sólo el primero. "
                     "Ningún valor de la matriz fue modificado para armar este archivo.")
nota.font = Font(name=F, size=9, italic=True, color="808080")
nota.alignment = Alignment(wrap_text=True, vertical="top")
h3.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=10)
h3.row_dimensions[r + 1].height = 58

h3.freeze_panes = "C3"
h3.print_title_rows = "1:2"
h3.print_area = "A1:J%d" % (r + 1)
h3.page_setup.orientation = "landscape"
h3.page_setup.paperSize = h3.PAPERSIZE_A4
h3.page_setup.fitToWidth = 1
h3.page_setup.fitToHeight = 0
h3.sheet_properties.pageSetUpPr.fitToPage = True

wb.active = 0
wb.save(OUT)
print("OK ->", OUT)
