#!/usr/bin/env python3
"""
mapa-xlsx — convierte .maestro/mapa-app.json en el Excel de trabajo.

POR QUE EXISTE
Decision de Enrique, 12-ago-2026: mapear las 187 pantallas antes de simplificar
la app. El HTML es para VER; este Excel es para CONFIGURAR, o sea para que el
criterio de Enrique quede escrito al lado de los numeros medidos.

REGLA DE DISENO
Las columnas medidas van bloqueadas visualmente (fondo gris, texto negro) y las
de criterio van en amarillo, que es la convencion de "aqui escribes tu". Nadie
tiene que adivinar donde meter mano.

USO
    python scripts/mapa-xlsx.py
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
ORIGEN = os.path.join(RAIZ, ".maestro", "mapa-app.json")
DESTINO = os.path.join(RAIZ, "R and D", "MAPA_APP_ATP.xlsx")

with open(ORIGEN, encoding="utf-8") as fh:
    datos = json.load(fh)

filas = datos["filas"]

# ── Paleta ───────────────────────────────────────────────────────────────────
GRIS = PatternFill("solid", fgColor="EEEEEE")
AMARILLO = PatternFill("solid", fgColor="FFF2CC")
ENCABEZADO = PatternFill("solid", fgColor="1F2933")
ROJO_SUAVE = PatternFill("solid", fgColor="FCE4E4")
VERDE_SUAVE = PatternFill("solid", fgColor="E4F5E4")
BORDE = Border(*[Side(style="thin", color="D0D0D0")] * 4)

FUENTE = "Arial"

COLUMNAS = [
    # (encabezado, ancho, tipo)  tipo: "medido" | "criterio"
    ("Ruta", 34, "medido"),
    ("Seccion", 14, "medido"),
    ("Toques desde tab", 15, "medido"),
    ("Enlazada desde", 9, "medido"),
    ("Enlaza a", 9, "medido"),
    ("Lineas", 8, "medido"),
    ("Colores clavados (criticos)", 14, "medido"),
    ("Cap. oscuro KB", 12, "medido"),
    ("Cap. claro KB", 12, "medido"),
    ("Liberada a produccion", 16, "criterio"),
    ("Tema OK", 11, "criterio"),
    ("Funcion OK", 11, "criterio"),
    ("Accesibilidad", 13, "criterio"),
    ("Bugs", 30, "criterio"),
    ("Diseno", 30, "criterio"),
    ("Notas", 40, "criterio"),
    ("Archivo", 40, "medido"),
]

wb = Workbook()

# ═══ Hoja 1 · LEE ESTO ═══════════════════════════════════════════════════════
guia = wb.active
guia.title = "Lee esto"
guia.column_dimensions["A"].width = 100

lineas_guia = [
    ("MAPA DE LA APP ATP", 16, True),
    ("", 11, False),
    (f"Generado el {datos['generado'][:10]} · {len(filas)} pantallas", 11, False),
    ("", 11, False),
    ("PARA QUE SIRVE", 12, True),
    ("Ver las 187 pantallas juntas para decidir que se sube de nivel, que se agrupa", 11, False),
    ("y que se esconde. El objetivo no es borrar: es dejar de exigirle al usuario que", 11, False),
    ("encuentre las cosas.", 11, False),
    ("", 11, False),
    ("DONDE ESCRIBES TU", 12, True),
    ("Las celdas AMARILLAS son tuyas. Las GRISES estan medidas del codigo y no", 11, False),
    ("hay que tocarlas: se regeneran con  node scripts/mapa-app.js", 11, False),
    ("", 11, False),
    ("QUE SIGNIFICA CADA COLUMNA MEDIDA", 12, True),
    ("Toques desde tab   Minimo de saltos desde uno de los 5 tabs. 'sin camino' NO", 11, False),
    ("                   quiere decir inalcanzable: quiere decir que el enlace no esta", 11, False),
    ("                   escrito en el archivo de la pantalla, sino en un componente.", 11, False),
    ("                   Es una cota, no una sentencia.", 11, False),
    ("Colores clavados   Literales de color en fondo, borde o texto. Son los que NO", 11, False),
    ("                   voltean con el tema. Mas alto = mas roto se ve en claro.", 11, False),
    ("Cap. KB            Peso de la captura. Muy bajo suele ser pantalla que no pinto.", 11, False),
    ("", 11, False),
    ("LAS COLUMNAS DE CRITERIO", 12, True),
    ("Liberada a produccion   Si / No / Oculta a proposito", 11, False),
    ("Tema OK                 Si / No / Solo oscuro / Solo claro", 11, False),
    ("Funcion OK              Si / No / A medias", 11, False),
    ("Accesibilidad           Si / No / Revisar", 11, False),
    ("Bugs, Diseno, Notas     Texto libre", 11, False),
    ("", 11, False),
    ("EJEMPLO DE COMO SE LLENA (fila 2 de la hoja Pantallas)", 12, True),
    ("La primera fila de datos viene llena como muestra. Borrala o sobreescribela.", 11, False),
]
for i, (texto, tam, negrita) in enumerate(lineas_guia, start=1):
    c = guia.cell(row=i, column=1, value=texto)
    c.font = Font(name=FUENTE, size=tam, bold=negrita)
    c.alignment = Alignment(wrap_text=False, vertical="center")

# ═══ Hoja 2 · Pantallas ══════════════════════════════════════════════════════
ws = wb.create_sheet("Pantallas")

for idx, (titulo, ancho, tipo) in enumerate(COLUMNAS, start=1):
    c = ws.cell(row=1, column=idx, value=titulo)
    c.font = Font(name=FUENTE, size=10, bold=True, color="FFFFFF")
    c.fill = ENCABEZADO
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = BORDE
    ws.column_dimensions[get_column_letter(idx)].width = ancho
ws.row_dimensions[1].height = 34

def kb(f, tema):
    cap = f["capturas"].get(tema)
    return cap["kb"] if cap and cap["existe"] else None

EJEMPLO = {
    "liberada": "Si",
    "tema": "Solo oscuro",
    "funcion": "Si",
    "accesibilidad": "Revisar",
    "bugs": "Las cards no cargan la imagen de fondo",
    "diseno": "Titulos truncados con espacio de sobra",
    "notas": "Ejemplo de como llenar. Borra o sobreescribe esta fila.",
}

r = 2
for i, f in enumerate(filas):
    prof = f["profundidad"]
    valores = [
        f["ruta"],
        f["seccion"],
        "sin camino" if prof is None else prof,
        len(f["enlazadaDesde"]),
        len(f["enlazaA"]),
        f["lineas"],
        f["coloresCriticos"],
        kb(f, "oscuro"),
        kb(f, "claro"),
    ]
    # La primera fila lleva el ejemplo; el resto va vacio.
    crit = EJEMPLO if i == 0 else {k: "" for k in EJEMPLO}
    valores += [crit["liberada"], crit["tema"], crit["funcion"],
                crit["accesibilidad"], crit["bugs"], crit["diseno"], crit["notas"]]
    valores.append(f["archivo"])

    for j, v in enumerate(valores, start=1):
        c = ws.cell(row=r, column=j, value=v)
        c.font = Font(name=FUENTE, size=10)
        c.border = BORDE
        c.alignment = Alignment(vertical="center",
                                wrap_text=COLUMNAS[j - 1][0] in ("Bugs", "Diseno", "Notas"))
        c.fill = AMARILLO if COLUMNAS[j - 1][2] == "criterio" else GRIS

    # Señales visuales sobre lo medido, para que el ojo priorice sin filtrar.
    if f["coloresCriticos"] >= 15:
        ws.cell(row=r, column=7).fill = ROJO_SUAVE
    if prof is not None and prof <= 2:
        ws.cell(row=r, column=3).fill = VERDE_SUAVE
    elif prof is not None and prof >= 4:
        ws.cell(row=r, column=3).fill = ROJO_SUAVE
    r += 1

ultima = r - 1
ws.freeze_panes = "B2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNAS))}{ultima}"

# Listas desplegables en las columnas de criterio cerradas.
for col, opciones in (
    ("J", '"Si,No,Oculta a proposito"'),
    ("K", '"Si,No,Solo oscuro,Solo claro"'),
    ("L", '"Si,No,A medias"'),
    ("M", '"Si,No,Revisar"'),
):
    dv = DataValidation(type="list", formula1=opciones, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col}2:{col}{ultima}")

# ═══ Hoja 3 · Resumen (con formulas, no numeros calculados en Python) ════════
res = wb.create_sheet("Resumen")
res.column_dimensions["A"].width = 42
res.column_dimensions["B"].width = 14

def bloque(fila, titulo):
    c = res.cell(row=fila, column=1, value=titulo)
    c.font = Font(name=FUENTE, size=12, bold=True)

bloque(1, "RESUMEN DEL MAPA")
renglones = [
    ("Pantallas totales", f"=COUNTA(Pantallas!A2:A{ultima})"),
    ("A 2 toques o menos de un tab", f'=COUNTIFS(Pantallas!C2:C{ultima},"<=2")'),
    ("A 3 toques o mas", f'=COUNTIFS(Pantallas!C2:C{ultima},">=3")'),
    ("Sin camino trazable", f'=COUNTIF(Pantallas!C2:C{ultima},"sin camino")'),
    ("Colores clavados criticos (suma)", f"=SUM(Pantallas!G2:G{ultima})"),
    ("Pantallas con 15+ colores clavados", f'=COUNTIF(Pantallas!G2:G{ultima},">=15")'),
    ("", ""),
    ("--- Lo que tu vayas llenando ---", ""),
    ("Liberadas a produccion", f'=COUNTIF(Pantallas!J2:J{ultima},"Si")'),
    ("NO liberadas", f'=COUNTIF(Pantallas!J2:J{ultima},"No")'),
    ("Ocultas a proposito", f'=COUNTIF(Pantallas!J2:J{ultima},"Oculta a proposito")'),
    ("Sin evaluar todavia", f'=COUNTBLANK(Pantallas!J2:J{ultima})'),
    ("Con tema roto", f'=COUNTIFS(Pantallas!K2:K{ultima},"No")+COUNTIFS(Pantallas!K2:K{ultima},"Solo oscuro")+COUNTIFS(Pantallas!K2:K{ultima},"Solo claro")'),
    ("Con funcion rota o a medias", f'=COUNTIFS(Pantallas!L2:L{ultima},"No")+COUNTIFS(Pantallas!L2:L{ultima},"A medias")'),
]
fila = 3
for etiqueta, formula in renglones:
    res.cell(row=fila, column=1, value=etiqueta).font = Font(name=FUENTE, size=10,
                                                            bold=etiqueta.startswith("---"))
    if formula:
        c = res.cell(row=fila, column=2, value=formula)
        c.font = Font(name=FUENTE, size=10, bold=True)
        c.fill = GRIS
        c.border = BORDE
    fila += 1

nota = res.cell(row=fila + 1, column=1,
                value="Las cuentas de criterio salen de la hoja Pantallas y se actualizan solas.")
nota.font = Font(name=FUENTE, size=9, italic=True)

# ═══ Hoja 4 · Por seccion ════════════════════════════════════════════════════
sec = wb.create_sheet("Por seccion")
secciones = sorted({f["seccion"] for f in filas})
for j, t in enumerate(["Seccion", "Pantallas", "Colores criticos", "A 3+ toques"], start=1):
    c = sec.cell(row=1, column=j, value=t)
    c.font = Font(name=FUENTE, size=10, bold=True, color="FFFFFF")
    c.fill = ENCABEZADO
    c.border = BORDE
sec.column_dimensions["A"].width = 24
for j in "BCD":
    sec.column_dimensions[j].width = 16

for i, s in enumerate(secciones, start=2):
    sec.cell(row=i, column=1, value=s).font = Font(name=FUENTE, size=10)
    sec.cell(row=i, column=2, value=f'=COUNTIF(Pantallas!B2:B{ultima},A{i})')
    sec.cell(row=i, column=3, value=f'=SUMIF(Pantallas!B2:B{ultima},A{i},Pantallas!G2:G{ultima})')
    sec.cell(row=i, column=4, value=f'=COUNTIFS(Pantallas!B2:B{ultima},A{i},Pantallas!C2:C{ultima},">=3")')
    for j in (2, 3, 4):
        c = sec.cell(row=i, column=j)
        c.font = Font(name=FUENTE, size=10)
        c.border = BORDE
        c.fill = GRIS
sec.auto_filter.ref = f"A1:D{len(secciones) + 1}"

os.makedirs(os.path.dirname(DESTINO), exist_ok=True)
wb.save(DESTINO)
print(f"  escrito: {os.path.relpath(DESTINO, RAIZ)}")
print(f"  {len(filas)} pantallas, {len(secciones)} secciones")
